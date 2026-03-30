from pylogix import PLC
from datetime import datetime
import time
import os
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------------- USER SETTINGS ----------------
PLC_IP = '192.168.1.10'
X_BASE_TAG = 'WestGantryXPositions'
Z_BASE_TAG = 'WestGantryZPositions'

START_INDEX = 0
END_INDEX = 10
SCAN_TIME = 0.25
DEADBAND = 0.001
SAVE_INTERVAL = 5   # seconds

LOG_FILE = 'WestGantry_AxisLog.xlsx'
SHEET_NAME = 'AxisChanges'
TABLE_NAME = 'GantryAxisLog'
# ----------------------------------------------

# ---------------- PLC ----------------
plc = PLC()
plc.IPAddress = PLC_IP
plc.ProcessorSlot = 0

# ---------------- TRACK LAST VALUES ----------------
last_values = {
    'X': {i: None for i in range(START_INDEX, END_INDEX + 1)},
    'Z': {i: None for i in range(START_INDEX, END_INDEX + 1)}
}

# ---------------- EXCEL SETUP ----------------
if not os.path.isfile(LOG_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    ws.append(['Date', 'Time', 'Axis', 'OldPreset', 'NewPreset'])

    table = Table(displayName=TABLE_NAME, ref='A1:E1')
    style = TableStyleInfo(name='TableStyleMedium9', showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(LOG_FILE)

# Load once
wb = load_workbook(LOG_FILE)
ws = wb[SHEET_NAME]

print("Monitoring X and Z presets (LOG ONLY ON CHANGE)...")

last_save_time = time.time()

# ---------------- MAIN LOOP ----------------
while True:
    try:
        for i in range(START_INDEX, END_INDEX + 1):
            for axis, base_tag in [('X', X_BASE_TAG), ('Z', Z_BASE_TAG)]:

                r = plc.Read(f'{base_tag}[{i}].Position.Preset')

                if r.Status != 'Success':
                    continue

                new_value = float(r.Value)

                if last_values[axis][i] is None:
                    last_values[axis][i] = new_value
                    continue

                old_value = last_values[axis][i]

                if abs(new_value - old_value) > DEADBAND:
                    now = datetime.now()

                    ws.append([
                        now.date(),
                        now,
                        axis,
                        old_value,
                        new_value
                    ])

                    ws.cell(ws.max_row, 1).number_format = 'yyyy-mm-dd'
                    ws.cell(ws.max_row, 2).number_format = 'hh:mm:ss.000 AM/PM'

                    last_values[axis][i] = new_value

                    print(f'Axis {axis}: {old_value} → {new_value}')

        # Resize table
        ws.tables[TABLE_NAME].ref = f'A1:E{ws.max_row}'

        # Save periodically (not every loop)
        if time.time() - last_save_time > SAVE_INTERVAL:
            wb.save(LOG_FILE)
            last_save_time = time.time()

        time.sleep(SCAN_TIME)

    except Exception as e:
        print(f'Logger error: {e}')
        time.sleep(2)